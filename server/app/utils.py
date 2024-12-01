from datetime import datetime
from flask import jsonify
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font,  NamedStyle
from openpyxl import load_workbook
import os
import logging
import subprocess

def format_date(date_str):
    try:
        # Attempt to parse full date-time with timezone
        date_obj = datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S %Z")
        return date_obj.strftime("%B %d, %Y")
    except ValueError:
        try:
            # Strip timezone manually and parse
            date_str = date_str.split(" GMT")[0]  # Remove GMT or similar
            date_obj = datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S")
            return date_obj.strftime("%B %d, %Y")
        except ValueError:
            try:
                # Try an alternative format: "YYYY-MM-DD"
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                return date_obj.strftime("%B %d, %Y")
            except ValueError:
                # If all parsing fails, return the original string
                return date_str


# Define a function to create a currency style with prefix
def create_currency_style(prefix):
    return NamedStyle(
        name=f"{prefix}_Currency",
        number_format=f'"{prefix}" @'
    )


# Create styles for different currencies
currency_styles = {
    "IDR": create_currency_style("Rp"),
    "SGD": create_currency_style("S$"),
    "USD": create_currency_style("$"),
}

# Default to USD if the currency type is not found


def get_currency_style(currency_type):
    return currency_styles.get(currency_type, currency_styles["USD"])


def generate_quotation_file(data, quotation_code, quotation_type):
    template_dir = "static/templates"
    output_dir = "static/xlsx"
    output_pdf_dir = "static/pdf"
    
    
    
    
    template_file = os.path.join(
        template_dir, "VMS.xlsx" if quotation_type == "VMS" else "Project.xlsx"
    )

    output_file = f"{quotation_code.replace('/', '_')}.xlsx"
    output_path = os.path.join(output_dir, output_file)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook = load_workbook(template_file)
    sheet = workbook.active

    # Insert customer and quotation details
    sheet["A7"] = data.get("quotation_customer", "")
    sheet["A8"] = data.get("quotation_email", "")  # Insert email into A8
    sheet["A10"] = data.get("quotation_customeraddress", "")
    sheet["A9"] = data.get("quotation_pt", "")
    sheet["F7"] = data.get("quotation_solution", "")  # Insert solution into F7

    # Format the date values to "Month DD, YYYY" format

    created_date = format_date(data.get("quotation_createdate", ""))

    valid_date = format_date(data.get("quotation_validdate", ""))

    print(created_date)
    print(valid_date)

    if quotation_type in ["BR", "PR"]:
        sheet["G1"] = created_date
        sheet["G2"] = valid_date
        sheet["G3"] = quotation_code
    else:
        sheet["H1"] = created_date
        sheet["H2"] = valid_date
        sheet["H3"] = quotation_code

    # Define border style (outline)
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    
    alignment_left = Alignment(horizontal="left", vertical="center")
    alignment_right = Alignment(horizontal="right")

    expenses = data.get("expenses", [])
    start_row = 15  # Starting row for expenses
    
    currency_type = data.get("currency_type", "USD")
    currency_style = get_currency_style(currency_type)

    for i, expense in enumerate(expenses):
        try:
            # Insert data into the respective columns
            sheet[f"A{start_row}"] = i + 1  # No column
            sheet[f"A{start_row}"].font = Font(name="Calibri")
            sheet[f"B{start_row}"] = expense.get("expense_description", "")
            sheet[f"B{start_row}"].font = Font(name="Calibri")
          

            # Merge cells and add borders
            if quotation_type in ["BR", "PR"]:
                sheet.merge_cells(f"A{start_row}:A{start_row + 1}")
                sheet.merge_cells(f"B{start_row}:E{start_row + 1}")
                sheet.merge_cells(f"F{start_row}:F{start_row + 1}")
                sheet.merge_cells(f"G{start_row}:G{start_row + 1}")
                
                
                sheet[f"F{start_row}"] = expense.get("expense_total", 0)
                sheet[f"F{start_row}"].style = currency_style
                sheet[f"F{start_row}"].font = Font(name="Calibri")
                sheet[f"F{start_row}"].alignment = alignment_right
                
                
                sheet[f"G{start_row}"] = expense.get("expense_linecount", 0)
                sheet[f"G{start_row}"].style = currency_style
                sheet[f"G{start_row}"].font = Font(name="Calibri", bold=True)
                sheet[f"G{start_row}"].alignment = alignment_right
                

                for row in range(start_row, start_row + 2):
                    for col in ["A", "B", "C", "D", "E", "F", "G"]:
                        sheet[f"{col}{row}"].border = thin_border
            else:
                sheet.merge_cells(f"A{start_row}:A{start_row + 1}")
                sheet.merge_cells(f"B{start_row}:E{start_row + 1}")
                sheet.merge_cells(f"F{start_row}:F{start_row + 1}")
                sheet.merge_cells(f"G{start_row}:G{start_row + 1}")
                sheet.merge_cells(f"H{start_row}:H{start_row + 1}")

                sheet[f"F{start_row}"] = expense.get("expense_total", 0)
                sheet[f"F{start_row}"].style = currency_style
                sheet[f"F{start_row}"].font = Font(name="Calibri")
                sheet[f"F{start_row}"].alignment = alignment_right
                
                
                
                sheet[f"G{start_row}"] = expense.get("expense_discount", 0)
                sheet[f"G{start_row}"].font = Font(name="Calibri")
                sheet[f"G{start_row}"].alignment = Alignment(horizontal="center")
                
                sheet[f"H{start_row}"] = expense.get("expense_linecount", 0)
                sheet[f"H{start_row}"].style = currency_style
                sheet[f"H{start_row}"].font = Font(name="Calibri", bold=True)
                sheet[f"H{start_row}"].alignment = alignment_right
                
                
                
                for row in range(start_row, start_row + 2):
                    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                        sheet[f"{col}{row}"].border = thin_border

            for col in ["A", "B", "F", "G", "H"]:
                for row in range(start_row, start_row + 2):
                    if col in ["F", "H"]:
                        sheet[f"{col}{row}"].alignment = Alignment(
                            horizontal="right", vertical="center"
                        )
                    else:
                        sheet[f"{col}{row}"].alignment = Alignment(
                            horizontal="left", vertical="center"
                        )

            start_row += 2
        except Exception as e:
            print(f"Error processing expense {i + 1}: {e}")
            
    vat_rate = data.get("vat_rate", "11%")  # Default to "11%" if vat_rate is missing
    vat_rate_label = f"Vat {vat_rate}%"  # Create the label dynamically

    if quotation_type == "VMS":
        summary_rows = [
            ("Subtotal", data.get("quotation_subtotal", 0), "FFBFBFBF"),
            ("Management Fee 5%", data.get("quotation_managementfee", 0), "FFBFBFBF"),
            (vat_rate_label, data.get("quotation_ppn", 0), "FFD8D8D8"),
            ("Grand Total", data.get("quotation_grandtotal", 0), "FF01205E"),
        ]

    elif quotation_type in ["BR", "PR"]:
        summary_rows = [
            ("Subtotal", data.get("quotation_subtotal", 0), "FFBFBFBF"),
            ("Vat 11%", data.get("quotation_ppn", 0), "FFD8D8D8"),
            ("Grand Total", data.get("quotation_grandtotal", 0), "FF01205E"),
        ]

    # Insert summary rows into the sheet
    for i, (field_name, field_value, color) in enumerate(summary_rows):
       

        # Merge cells for label column (A to G for BR/PR, A to H for VMS)
        if quotation_type in ["BR", "PR"]:
            # Merge A-F for label
            for row in range(start_row, start_row + 1):
                for col in range(1, 7):  # Columns A to F
                    sheet.cell(row=row, column=col).border = thin_border
            sheet.merge_cells(f"A{start_row}:F{start_row}")
            value_column = "G"  # Value in column G
        else:
            for row in range(start_row, start_row + 1):
                for col in range(1, 8):  # Columns A to F
                    sheet.cell(row=row, column=col).border = thin_border
            sheet.merge_cells(f"A{start_row}:G{start_row}")
            value_column = "H"  # Value in column H

        # Set the label cell value and style
        label_cell = sheet[f"A{start_row}"]
        label_cell.value = field_name
        label_cell.fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid")
        label_cell.alignment = alignment_right
        label_cell.font = Font(
            name="Calibri",  bold=True, color="FFFFFF" if field_name == "Grand Total" else "000000"
        )
        label_cell.border = thin_border

        # Set the value cell with the appropriate style
        value_cell = sheet[f"{value_column}{start_row}"]
        value_cell.value = field_value
        value_cell.style = currency_style
        value_cell.alignment = alignment_right
        value_cell.font = Font(name="Calibri", bold=True)
        value_cell.border = thin_border
       
        
        start_row += 1

   # Outline the cells and merge them
    if quotation_type in ["BR", "PR"]:
        for col in range(1, 8):  # Columns A (1) to G (7)
            cell = sheet.cell(row=start_row, column=col)
            cell.border = Border(top=thin_border.top, bottom=thin_border.bottom)  # Apply top and bottom borders only
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        for col in range(1, 9):  # Columns A (1) to H (8)
            cell = sheet.cell(row=start_row, column=col)
            cell.border = Border(top=thin_border.top, bottom=thin_border.bottom)  # Apply top and bottom borders only
        sheet.merge_cells(f"A{start_row}:H{start_row}")


    # Insert Terms and Conditions (TOS)
    tos_data = data.get("terms_and_conditions", [])
    start_row += 1  # Move to the row below the summary section

    # Determine the merge range based on quotation type
    if quotation_type in ["BR", "PR"]:
        merge_columns = "G"
        last_col = 7  # Column G
    else:
        merge_columns = "H"
        last_col = 8  # Column H

    # Add header for Terms and Conditions
    for col in range(1, last_col + 1):  # Apply styling before merging
        cell = sheet.cell(row=start_row, column=col)
        cell.value = "Terms and Conditions" if col == 1 else None
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(
            start_color="FF01205E", end_color="FF01205E", fill_type="solid"
        )
        cell.border = thin_border

    sheet.merge_cells(f"A{start_row}:{merge_columns}{start_row}")  # Merge cells for the header

    start_row += 1  # Move to the first TOS row

    # Add TOS items
    for i, tos in enumerate(tos_data):
        tos_row = start_row + i

        # Add TOS number (Column A) and apply borders
        tos_number_cell = sheet.cell(row=tos_row, column=1)
        tos_number_cell.value = i + 1
        tos_number_cell.alignment = Alignment(
            horizontal="center", vertical="center")
        tos_number_cell.border = Border(
            left=thin_border.left,
            top=thin_border.top,
            bottom=thin_border.bottom
        )

        # Apply styling to each cell before merging for the description
        for col in range(2, last_col + 1):  # Columns B to G or B to H
            cell = sheet.cell(row=tos_row, column=col)
            if col == 2:  # First column of the description
                cell.value = tos.get("tos_description", "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Apply borders for each cell
            if col == last_col:  # Last column gets the right border
                cell.border = Border(
                    top=thin_border.top,
                    bottom=thin_border.bottom,
                    right=thin_border.right
                )
            else:  # Other columns get top and bottom borders only
                cell.border = Border(
                    top=thin_border.top,
                    bottom=thin_border.bottom
                )

        # Merge description cells after applying styles and borders
        sheet.merge_cells(f"B{tos_row}:{merge_columns}{tos_row}")

    start_row += len(tos_data) + 2  # Adjust for TOS rows and spacing

    # Add "Sincerely,"
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
        
    sheet[f"A{start_row}"].value = "Sincerely,"
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="left", vertical="center")
    sheet[f"A{start_row}"].font = Font( name="Arial")
    

    # Add PIC name
    start_row += 2
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = data.get("quotation_pic", "")
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="left", vertical="center")
    sheet[f"A{start_row}"].font = Font(bold=True, name="Arial")

    # Add company name
    start_row += 1
    sheet.merge_cells(f"A{start_row}:C{start_row}")
    sheet[f"A{start_row}"].value = "PT Moonlay Technologies"
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="left", vertical="center")
    sheet[f"A{start_row}"].font = Font(name="Arial")

    # Add "Thank you for your business!"
    start_row += 4
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = "Thank you for your business!"
    sheet[f"A{start_row}"].font = Font(bold=True, name="Arial" )
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="center", vertical="center")
    
    # Add contact details
    start_row += 1
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = "Should you have any enquiries concerning this quote, please contact Murni on +62-823-6506-0216"
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="center", vertical="center")
    sheet[f"A{start_row}"].font = Font(name="Arial")

    start_row += 1
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = "Equity Tower 25th Floor, Suite H, SCBD Lot. 9, Jln. Jenderal Sudirman Kav. 52-53 "
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="center", vertical="center")
    sheet[f"A{start_row}"].font = Font(name="Arial")

    start_row += 1
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = "Kel. Senayan, Kec. Kebayoran Baru, Kota Administrasi Jakarta Selatan, Prov. DKI Jakarta 12190 "
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="center", vertical="center")
    sheet[f"A{start_row}"].font = Font(name="Arial")

    start_row += 1
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = "Tel: (+62-21) 290-350-89 | Fax: (+62-21) 290-350-88 | E-mail: murni.telaumbanua@moonlay.com | Web: www.moonlay.com"
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="center", vertical="center")
    sheet[f"A{start_row}"].font = Font(name="Arial")

    # Add "HIGHLY CONFIDENTIAL"
    start_row += 2
    if quotation_type in ["BR", "PR"]:
        sheet.merge_cells(f"A{start_row}:G{start_row}")
    else:
        sheet.merge_cells(f"A{start_row}:H{start_row}")
    sheet[f"A{start_row}"].value = "HIGHLY CONFIDENTIAL"
    sheet[f"A{start_row}"].font = Font(bold=True, color="FF0000", name="Arial")
    sheet[f"A{start_row}"].alignment = Alignment(
        horizontal="center", vertical="center")

    # Save the updated Excel file
    workbook.save(output_path)
    
    if not os.path.exists(output_pdf_dir):
        os.makedirs(output_pdf_dir)
        
    print(output_path)

    # Generate absolute paths
    pdf_file_name = f"{quotation_code.replace('/', '_')}.pdf"
    pdf_output_path = os.path.join(output_pdf_dir, pdf_file_name)
    

    print("Source file path:", output_path)
    print("Output directory path:", pdf_output_path)
    

    try:
        result = subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                os.path.dirname(pdf_output_path),
                os.path.abspath(output_path)
            ],
            check=True,
            capture_output=True,
            text=True
        )
        print("LibreOffice Output:", result.stdout)
        print("LibreOffice Errors:", result.stderr)
    except subprocess.CalledProcessError as e:
        print(f"Error: {e.stderr}")
        return jsonify({'error': 'Conversion to PDF failed', 'details': e.stderr}), 500


    # Check if the PDF was created
    return jsonify({
        'xlsxUrl': f'/static/xlsx/{output_file}',
        'pdfUrl': f'/static/pdf/{quotation_code.replace("/", "_")}.pdf'
    }), 200
